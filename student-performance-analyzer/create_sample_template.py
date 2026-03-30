# Utility script — run once to generate templates/sample_import.xlsx

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Students"

headers = [
    'student_code', 'first_name', 'last_name', 'email',
    'subject', 'assessment_name', 'max_score', 'score', 'topic', 'date',
]

# Style the header row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[cell.column_letter].width = 20

# Sample data rows
rows = [
    ['ALU-2026-001', 'Alice',   'Mwangi',  'alice@alu.edu',   'Mathematics', 'Quiz 1',  100, 85,  'Algebra',    '2026-03-01'],
    ['ALU-2026-002', 'Brian',   'Osei',    'brian@alu.edu',   'Mathematics', 'Quiz 1',  100, 52,  'Algebra',    '2026-03-01'],
    ['ALU-2026-003', 'Clara',   'Diallo',  'clara@alu.edu',   'Mathematics', 'Quiz 1',  100, 91,  'Algebra',    '2026-03-01'],
    ['ALU-2026-001', 'Alice',   'Mwangi',  'alice@alu.edu',   'Physics',     'Midterm', 100, 70,  'Mechanics',  '2026-03-05'],
    ['ALU-2026-002', 'Brian',   'Osei',    'brian@alu.edu',   'Physics',     'Midterm', 100, 35,  'Mechanics',  '2026-03-05'],
]

for row in rows:
    ws.append(row)

wb.save("templates/sample_import.xlsx")
print("templates/sample_import.xlsx created.")
