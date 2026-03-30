# services/data_importer.py — Bulk .xlsx import with validation and batch insert

import os
import openpyxl
from utils.validators import (
    validate_student_code, validate_email,
    validate_non_empty, validate_score,
)
from utils.display_helpers import progress_bar, print_success, print_error, print_warning

REQUIRED_COLUMNS = {
    'student_code', 'first_name', 'last_name', 'email',
    'subject', 'assessment_name', 'max_score', 'score', 'date',
}


class DataImporter:
    """Parses .xlsx files and batch-inserts student/score records into the database.

    Expected column headers (topic is optional):
        student_code | first_name | last_name | email | subject |
        assessment_name | max_score | score | topic | date
    """

    def __init__(self, db_manager):
        self.db = db_manager

    def import_from_xlsx(self, file_path):
        """Main entry point. Returns (imported, skipped, errors) counts."""
        imported = skipped = errors = 0
        error_log = []

        # --- Validate file ---
        if not os.path.exists(file_path):
            print_error(f"File not found: {file_path}")
            return 0, 0, 1

        if not file_path.endswith('.xlsx'):
            print_error("File must be a .xlsx spreadsheet.")
            return 0, 0, 1

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # --- Validate headers ---
        raw_headers = [cell.value for cell in sheet[1]]
        headers = [str(h).strip().lower() if h else '' for h in raw_headers]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            print_error(f"Missing required columns: {', '.join(sorted(missing))}")
            return 0, 0, 1

        total_rows = sheet.max_row - 1
        if total_rows == 0:
            print_warning("Spreadsheet is empty.")
            return 0, 0, 0

        # --- Process rows ---
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            progress_bar(idx, total_rows)
            data = dict(zip(headers, row))

            # Skip fully empty rows
            if all(v is None for v in data.values()):
                skipped += 1
                continue

            # Validate fields
            row_errors = self._validate_row(data, idx)
            if row_errors:
                error_log.extend(row_errors)
                errors += 1
                continue

            # Normalise values
            student_code = str(data['student_code']).strip().upper()
            first_name   = str(data['first_name']).strip()
            last_name    = str(data['last_name']).strip()
            email        = str(data['email']).strip()
            subject_name = str(data['subject']).strip()
            assess_name  = str(data['assessment_name']).strip()
            max_score    = float(data['max_score'])
            score_value  = float(data['score'])
            topic_name   = str(data.get('topic') or '').strip() or None
            date_given   = str(data['date']).strip() if data.get('date') else None

            try:
                # Student — get or create
                student = self.db.execute_query(
                    "SELECT student_id FROM students WHERE student_code = %s",
                    (student_code,), fetch='one'
                )
                if not student:
                    sid = self.db.execute_query(
                        "INSERT INTO students (student_code, first_name, last_name, email) "
                        "VALUES (%s, %s, %s, %s)",
                        (student_code, first_name, last_name, email)
                    )
                    student_id = sid
                else:
                    student_id = student['student_id']

                # Subject — get or create
                subject = self.db.execute_query(
                    "SELECT subject_id FROM subjects WHERE subject_name = %s",
                    (subject_name,), fetch='one'
                )
                if not subject:
                    subj_id = self.db.execute_query(
                        "INSERT INTO subjects (subject_name) VALUES (%s)", (subject_name,)
                    )
                else:
                    subj_id = subject['subject_id']

                # Assessment — get or create
                assessment = self.db.execute_query(
                    "SELECT assessment_id FROM assessments "
                    "WHERE subject_id = %s AND assessment_name = %s",
                    (subj_id, assess_name), fetch='one'
                )
                if not assessment:
                    assess_id = self.db.execute_query(
                        "INSERT INTO assessments (subject_id, assessment_name, max_score, date_given) "
                        "VALUES (%s, %s, %s, %s)",
                        (subj_id, assess_name, max_score, date_given)
                    )
                else:
                    assess_id = assessment['assessment_id']

                # Topic — get or create (optional)
                topic_id = None
                if topic_name:
                    topic = self.db.execute_query(
                        "SELECT topic_id FROM topics WHERE subject_id = %s AND topic_name = %s",
                        (subj_id, topic_name), fetch='one'
                    )
                    if not topic:
                        topic_id = self.db.execute_query(
                            "INSERT INTO topics (subject_id, topic_name) VALUES (%s, %s)",
                            (subj_id, topic_name)
                        )
                    else:
                        topic_id = topic['topic_id']

                # Score — insert
                self.db.execute_query(
                    "INSERT INTO scores (student_id, assessment_id, score_value, topic_id) "
                    "VALUES (%s, %s, %s, %s)",
                    (student_id, assess_id, score_value, topic_id)
                )
                imported += 1

            except Exception as e:
                error_log.append(f"  Row {idx}: unexpected error — {e}")
                errors += 1

        # --- Summary ---
        print()  # newline after progress bar
        print_success(f"Imported : {imported}")
        if skipped:
            print_warning(f"Skipped  : {skipped}")
        if errors:
            print_error(f"Errors   : {errors}")
            for msg in error_log:
                print_error(msg)

        return imported, skipped, errors

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _validate_row(self, data, idx):
        """Return a list of error strings for this row (empty = valid)."""
        errs = []

        code = str(data.get('student_code') or '').strip().upper()
        valid, msg = validate_student_code(code)
        if not valid:
            errs.append(f"  Row {idx}: student_code — {msg}")

        email = str(data.get('email') or '').strip()
        valid, msg = validate_email(email)
        if not valid:
            errs.append(f"  Row {idx}: email — {msg}")

        for field in ('first_name', 'last_name', 'subject', 'assessment_name'):
            valid, msg = validate_non_empty(data.get(field), field)
            if not valid:
                errs.append(f"  Row {idx}: {field} — {msg}")

        try:
            max_score = float(data.get('max_score') or 0)
            score     = float(data.get('score') or -1)
            valid, msg = validate_score(score, max_score)
            if not valid:
                errs.append(f"  Row {idx}: score — {msg}")
        except (TypeError, ValueError):
            errs.append(f"  Row {idx}: max_score/score must be numbers")

        return errs
